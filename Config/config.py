import os
from openpyxl import load_workbook


class TestData:
    BASE_URL = "https://opensource-demo.orangehrmlive.com"
    USER_NAME = "Admin"
    PASSWORD = "admin123"
    HOMEPAGE_TITLE = "OrangeHRM"

    # Get the current directory of the project
    project_dir = os.getcwd()
    # Create the "Failed" folder
    failed_folder = os.path.join(project_dir, "Failed")
    if not os.path.exists(failed_folder):
        os.makedirs(failed_folder)

    def load_test_data():
        workbook = load_workbook(filename='D:\\SeleniumPOM\\Data\\SearchContents.xlsx')
        sheet = workbook.active
        test_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            Search_Query, Expected_Results = row
            test_data.append((Search_Query, Expected_Results))

        return test_data
